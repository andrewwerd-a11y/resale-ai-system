from __future__ import annotations
import csv
import json
from datetime import datetime
from pathlib import Path

from packages.domain.src.entities.item import Item
from packages.core.src.config import get_settings


COLUMNS = [
    "SKU", "Title", "Category", "Brand", "Type", "Department", "Size", "Color",
    "Material", "Style", "Condition", "Condition ID", "Condition Notes",
    "Author", "Format", "ISBN", "Franchise", "Character",
    "Features", "Defects", "Keywords",
    "Est. Price", "List Price", "Sold Price", "eBay Fees", "Net Profit",
    "AI Confidence", "AI Model", "Status",
    "eBay Listing ID", "eBay Listing URL",
    "Date Listed", "Date Sold", "Date Created",
]


def _fmt_list(lst: list) -> str:
    return "; ".join(lst) if lst else ""


def _fmt_date(dt) -> str:
    if dt is None:
        return ""
    if isinstance(dt, datetime):
        return dt.strftime("%Y-%m-%d")
    return str(dt)


def item_to_row(item: Item) -> dict:
    return {
        "SKU": item.sku,
        "Title": item.title or "",
        "Category": item.category or "",
        "Brand": item.brand or "",
        "Type": item.item_type or "",
        "Department": item.department or "",
        "Size": item.size or "",
        "Color": item.color or "",
        "Material": item.material or "",
        "Style": item.style or "",
        "Condition": item.condition or "",
        "Condition ID": item.condition_id or "",
        "Condition Notes": item.condition_notes or "",
        "Author": item.author or "",
        "Format": item.book_format or "",
        "ISBN": item.isbn or "",
        "Franchise": item.franchise or "",
        "Character": item.character or "",
        "Features": _fmt_list(item.features),
        "Defects": _fmt_list(item.defects),
        "Keywords": _fmt_list(item.keywords),
        "Est. Price": item.estimated_price or "",
        "List Price": item.list_price or "",
        "Sold Price": item.sold_price or "",
        "eBay Fees": item.ebay_fees or "",
        "Net Profit": item.net_profit or "",
        "AI Confidence": f"{item.ai_confidence:.2f}" if item.ai_confidence else "",
        "AI Model": item.ai_model or "",
        "Status": item.status,
        "eBay Listing ID": item.ebay_listing_id or "",
        "eBay Listing URL": item.ebay_listing_url or "",
        "Date Listed": _fmt_date(item.date_listed),
        "Date Sold": _fmt_date(item.date_sold),
        "Date Created": _fmt_date(item.date_created),
    }


def generate_csv(items: list[Item], output_path: Path | None = None) -> Path:
    settings = get_settings()
    if output_path is None:
        ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        output_path = settings.export_dir / f"master_inventory_{ts}.csv"

    output_path.parent.mkdir(parents=True, exist_ok=True)

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=COLUMNS)
        writer.writeheader()
        for item in items:
            writer.writerow(item_to_row(item))

    return output_path


def generate_excel(items: list[Item], output_path: Path | None = None) -> Path:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        # Fallback to CSV
        return generate_csv(items, output_path)

    settings = get_settings()
    if output_path is None:
        ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        output_path = settings.export_dir / f"master_inventory_{ts}.xlsx"

    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, item in enumerate(items, 2):
        row = item_to_row(item)
        for col_idx, col_name in enumerate(COLUMNS, 1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(col_name, ""))

    wb.save(output_path)
    return output_path
